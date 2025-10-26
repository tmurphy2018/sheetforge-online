from fastapi import FastAPI, Body, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Any, Dict, List, Optional
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- helpers ----------
THIN = Side(style="thin", color="D1D5DB")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="F8FAFC")

def px_to_col_width(px: int) -> float:
    # rough conversion: 1 Excel width ~ 7 px baseline, adjust slightly
    return max(3, round(px / 7.2, 1))

def write_table(ws, start_row, start_col, model: Dict[str, Any]):
    """
    model: { columns: [..], data: [[..]], name: str, formats: {colName: code}, cellFmt: {"r,c": code} }
    returns (end_row, end_col)
    """
    cols = model.get("columns") or []
    data = model.get("data") or []
    formats = model.get("formats") or {}
    cell_fmt = model.get("cellFmt") or {}

    # header row
    for j, col in enumerate(cols, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER_THIN
        cell.alignment = Alignment(vertical="center")
    r = start_row + 1

    # data rows
    for row in data:
        for j, col in enumerate(cols, start=start_col):
            idx = j - start_col
            val = row[idx] if idx < len(row) else ""
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER_THIN
            c.alignment = Alignment(vertical="top")
        r += 1

    # simple column widths guess
    for j, col in enumerate(cols, start=start_col):
        width = px_to_col_width(100 if col.lower() not in ("amount","total","value","price","cost") else 80)
        ws.column_dimensions[get_column_letter(j)].width = width

    # apply column formats
    for name, code in formats.items():
        if name in cols:
            j = start_col + cols.index(name)
            if code == "currency":
                for rr in range(start_row+1, r):
                    ws.cell(row=rr, column=j).number_format = '"$"#,##0.00'
            elif code == "number2":
                for rr in range(start_row+1, r):
                    ws.cell(row=rr, column=j).number_format = '#,##0.00'
            elif code == "date":
                for rr in range(start_row+1, r):
                    ws.cell(row=rr, column=j).number_format = 'mm/dd/yyyy'
            # text/upper/proper left as export-time content changes

    # apply per-cell overrides
    for key, code in cell_fmt.items():
        try:
            rr, cc = key.split(",")
            rr, cc = int(rr), int(cc)
        except:
            continue
        rr = start_row + rr  # convert local to absolute? keep simple: data rows start at +1
        cc = start_col + cc - 1
        cell = ws.cell(row=rr, column=cc)
        if code == "currency":
            cell.number_format = '"$"#,##0.00'
        elif code == "number2":
            cell.number_format = '#,##0.00'
        elif code == "date":
            cell.number_format = 'mm/dd/yyyy'
        # text transforms are visual; Excel has no "upper" format; would require values altered earlier.

    end_row = r - 1
    end_col = start_col + len(cols) - 1
    return end_row, end_col

def write_kpi(ws, start_row, start_col, model: Dict[str, Any]):
    title = model.get("title") or "KPI"
    sub = model.get("sub") or ""
    value = model.get("value") or 0
    # title
    ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+3)
    # value
    vcell = ws.cell(row=start_row+1, column=start_col, value=value)
    vcell.font = Font(bold=True, size=18)
    ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+2, end_column=start_col+3)
    # border around block
    for rr in range(start_row, start_row+3):
        for cc in range(start_col, start_col+4):
            ws.cell(row=rr, column=cc).border = BORDER_THIN
    # subtitle
    ws.cell(row=start_row+3, column=start_col, value=sub).font = Font(size=10, color="64748B")

def write_chart(ws, start_row, start_col, model: Dict[str, Any]):
    title = model.get("title") or "Chart"
    ctype = model.get("chartType") or "bar"
    # write a tiny table to chart
    labels = ["Jan","Feb","Mar","Apr","May","Jun"]
    values = [10, 22, 18, 30, 26, 35]
    ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True)
    for i, lab in enumerate(labels):
        ws.cell(row=start_row+1+i, column=start_col, value=lab)
        ws.cell(row=start_row+1+i, column=start_col+1, value=values[i]).number_format = '#,##0'
        ws.cell(row=start_row+1+i, column=start_col).border = BORDER_THIN
        ws.cell(row=start_row+1+i, column=start_col+1).border = BORDER_THIN
    # chart
    data = Reference(ws, min_col=start_col+1, min_row=start_row, max_col=start_col+1, max_row=start_row+len(labels))
    cats = Reference(ws, min_col=start_col, min_row=start_row+1, max_row=start_row+len(labels))
    chart = BarChart()
    chart.title = title
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"{get_column_letter(start_col+3)}{start_row}")

def set_page_setup(ws, settings: Dict[str, Any]):
    page = settings.get("page") or {}
    size = (page.get("size") or "Letter").lower()
    orient = (page.get("orientation") or "portrait").lower()

    # paper sizes mapping (Excel ids)
    PAPER = {"letter":1, "legal":5, "a4":9, "a3":8}
    ws.page_setup.paperSize = PAPER.get(size, 1)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT if orient=="portrait" else ws.ORIENTATION_LANDSCAPE

    margin = page.get("margin") or {}
    # convert px (~96dpi) to inches as rough estimate
    def px_to_in(px): return float(px)/96.0
    ws.page_margins.top = px_to_in(margin.get("top",40))
    ws.page_margins.bottom = px_to_in(margin.get("bottom",40))
    ws.page_margins.left = px_to_in(margin.get("left",40))
    ws.page_margins.right = px_to_in(margin.get("right",40))

# ---------- API ----------
class CompilePayload(BaseModel):
    project: str
    sheets: List[Dict[str, Any]]
    settings: Dict[str, Any]

@app.post("/api/compile")
def compile_workbook(payload: CompilePayload):
    wb = openpyxl.Workbook()
    # remove default sheet; we will add per our list
    default = wb.active
    wb.remove(default)

    for sidx, sheet in enumerate(payload.sheets):
        ws = wb.create_sheet(title=sheet.get("name","Sheet"+str(sidx+1))[:31])
        set_page_setup(ws, payload.settings)

        # Note: we map canvas units to cell grid rough positions.
        # top-left content origin is R2 C2 to leave a small border
        for w in sheet.get("widgets", []):
            x = int(w.get("x",1)); y = int(w.get("y",1)); col = x; row = y
            typ = w.get("type")
            # Title/sub at top (in Excel area)
            if typ == "table":
                title = w.get("title") or "Table"
                ws.cell(row=row-1, column=col, value=title).font = Font(bold=True)
                end_r, end_c = write_table(ws, row, col, w)
            elif typ == "kpi":
                write_kpi(ws, row, col, w)
            elif typ == "chart":
                write_chart(ws, row, col, w)
            else:
                # button or other -> simple label
                ws.cell(row=row, column=col, value=w.get("title") or "Button").border = BORDER_THIN

        # global niceties
        for r in range(1, 200):
            ws.row_dimensions[r].height = None  # let Excel auto-fit mostly

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(content=bio.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{payload.project.replace(" ","_")}.xlsx"'})
